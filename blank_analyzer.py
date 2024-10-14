import fitz  # PyMuPDF
import cv2
import numpy as np
import os
from tkinter import Tk, filedialog, messagebox, StringVar, ttk, Canvas
import threading
import subprocess
import platform
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import unicodedata
from concurrent.futures import ThreadPoolExecutor
from ttkthemes import ThemedTk
from PIL import Image, ImageEnhance, ImageFilter, ImageTk
import io
import sys

# Configuração do Tesseract OCR
tessdata_prefix = r'C:/Program Files/Tesseract-OCR/tessdata/'
os.environ['TESSDATA_PREFIX'] = tessdata_prefix

import pytesseract
pytesseract.pytesseract.tesseract_cmd = r'C:/Program Files/Tesseract-OCR/tesseract.exe'

def test_tesseract_setup():
    try:
        tessdata_prefix_env = os.environ.get('TESSDATA_PREFIX')
        print(f"TESSDATA_PREFIX: {tessdata_prefix_env}")
        if not tessdata_prefix_env or not os.path.isdir(tessdata_prefix_env):
            raise EnvironmentError("TESSDATA_PREFIX não está configurado corretamente.")
        tesseract_version = pytesseract.get_tesseract_version()
        print(f"Versão do Tesseract: {tesseract_version}")
        messagebox.showinfo("Tesseract OCR", f"Tesseract OCR instalado corretamente.\nVersão: {tesseract_version}")
    except Exception as e:
        print(f"Erro ao inicializar o Tesseract OCR: {e}")
        messagebox.showerror("Erro Tesseract OCR", f"Erro ao inicializar o Tesseract OCR.\n{str(e)}")
        sys.exit(1)

def sanitize_filename(filename):
    sanitized = unicodedata.normalize('NFKD', filename).encode('ascii', 'ignore').decode('ascii')
    print(f"Sanitized filename: {sanitized}")
    return sanitized

def is_blank_or_noisy(image, white_threshold=215, pixel_threshold=0.993):
    if image is None:
        print("Imagem é None")
        return False, 0

    white_pixel_percentage = np.mean(image > white_threshold)
    print(f"Porcentagem de pixels brancos: {white_pixel_percentage:.2%}")
    return ("branca sem necessidade de ocr" if white_pixel_percentage >= pixel_threshold else False), white_pixel_percentage

def perform_ocr_and_reclassify(image, min_text_length=2, is_blank=False):
    try:
        if is_blank:
            print("Aumentando resolução da imagem e aplicando filtros para OCR.")
            image = image.resize((int(image.width * 2), int(image.height * 2)), Image.LANCZOS)
            image = image.filter(ImageFilter.MedianFilter(size=3))
            image = ImageEnhance.Contrast(image).enhance(2.0)
            image = ImageEnhance.Sharpness(image).enhance(2.0)
        
        with io.BytesIO() as output:
            image.save(output, format="PNG", dpi=(600, 600))
            output.seek(0)
            with Image.open(output) as image_dpi:
                print("Aumentando contraste e nitidez da imagem para OCR.")
                image_bw = ImageEnhance.Sharpness(ImageEnhance.Contrast(image_dpi).enhance(2.0)).enhance(2.0)
                image_bw = image_bw.convert('L').point(lambda x: 0 if x < 128 else 255, '1')
                text = pytesseract.image_to_string(image_bw, lang='eng+por')
                print(f"Texto extraído pelo OCR: {text}")
        
        cleaned_text = ''.join(text.split())
        print(f"Texto limpo: {cleaned_text}")
        return (True, "precisa de revisão") if len(cleaned_text) > min_text_length else (False, "Sem Texto")
    except pytesseract.TesseractError as e:
        print(f"Erro durante o OCR: {e}")
        return False, "Erro OCR"

def display_image_on_canvas(image):
    """Exibe a imagem no centro do Canvas ajustando o tamanho do Canvas para a imagem e a centralizando."""
    # Obtém as dimensões do Canvas e da imagem
    canvas_width = canvas.winfo_width()
    canvas_height = canvas.winfo_height()
    img_width, img_height = image.size

    # Calcula a posição para centralizar a imagem no Canvas
    x = (canvas_width - img_width) // 2
    y = (canvas_height - img_height) // 2

    tk_image = ImageTk.PhotoImage(image)

    canvas.delete("all")
    canvas.create_image(x, y, anchor="nw", image=tk_image)
    canvas.image = tk_image  # Referência para evitar garbage collection

def analyze_single_pdf(pdf_path, ws, total_pages_processed, total_pages, progress_var, progress_label, progress_bar):
    print(f"Analisando PDF: {pdf_path}")
    with fitz.open(pdf_path) as pdf_document:
        pdf_name = sanitize_filename(os.path.basename(pdf_path))

        for page_num in range(pdf_document.page_count):
            try:
                print(f"Processando página {page_num + 1} de {pdf_document.page_count}")
                page = pdf_document.load_page(page_num)
                pix = page.get_pixmap()
                with Image.open(io.BytesIO(pix.tobytes("png"))) as img:
                    display_image_on_canvas(img)

                is_blank, white_pixel_percentage = is_blank_or_noisy(np.array(img.convert('L')))
                if is_blank == "branca sem necessidade de ocr":
                    status, ocr_text = "Branca sem necessidade de OCR", "Não"
                elif 0.97 <= white_pixel_percentage < 0.995:
                    relevant_text_found, text_or_revision = perform_ocr_and_reclassify(img, is_blank=True)
                    status, ocr_text = (text_or_revision if relevant_text_found else "Em branco após OCR", "Sim" if relevant_text_found else "Não")
                else:
                    status, ocr_text = "OK", "Não"

                print(f"Status da página {page_num + 1}: {status}, OCR realizado: {ocr_text}")
                ws.append([pdf_name, page_num + 1, status, f"{white_pixel_percentage:.2%}", ocr_text])

            except (FileNotFoundError, ValueError, IOError) as e:
                print(f"Erro ao processar a página {page_num + 1} do PDF {pdf_name}: {e}")
                ws.append([pdf_name, page_num + 1, f"Erro: {str(e)}", None, None])

            total_pages_processed[0] += 1
            progress_percentage = (total_pages_processed[0] / total_pages) * 100
            progress_var.set(progress_percentage)
            progress_label.config(text=f"Analisando... {int(float(progress_var.get()))}%")
            progress_bar.update_idletasks()

def select_directory():
    global directory
    directory = filedialog.askdirectory()
    if directory:
        select_label.config(text=f"Diretório Selecionado: {directory}")
        analyze_button.config(state="normal")

def start_analysis():
    if directory:
        analyze_button.config(state="disabled")
        threading.Thread(target=run_analysis_thread, args=(directory, progress_var, progress_label, progress_bar)).start()

def run_analysis_thread(directory, progress_var, progress_label, progress_bar):
    output_xlsx = os.path.join(directory, "analysis_report.xlsx")
    analyze_pdfs_in_directory(directory, progress_var, progress_label, output_xlsx, progress_bar)

def open_folder():
    if directory:
        if platform.system() == "Windows":
            os.startfile(directory)
        elif platform.system() == "Darwin":
            subprocess.Popen(["open", directory])
        elif platform.system() == "Linux":
            subprocess.Popen(["xdg-open", directory])

def analyze_pdfs_in_directory(pdf_dir, progress_var, progress_label, output_xlsx, progress_bar):
    wb = Workbook()
    ws = wb.active
    ws.title = "PDF Analysis Report"
    ws.append(["Arquivo PDF", "Página", "Status", "Porcentagem de Pixels Brancos", "Texto Extraído"])

    pdf_files = [os.path.join(pdf_dir, f) for f in os.listdir(pdf_dir) if f.lower().endswith('.pdf')]
    total_pages = sum(fitz.open(pdf_file).page_count for pdf_file in pdf_files)
    total_pages_processed = [0]

    with ThreadPoolExecutor(max_workers=4) as executor:
        futures = [executor.submit(analyze_single_pdf, pdf_file, ws, total_pages_processed, total_pages, progress_var, progress_label, progress_bar) for pdf_file in pdf_files]
        for future in futures:
            future.result()

    tab = Table(displayName="PDFAnalysisTable", ref=f"A1:E{ws.max_row}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showRowStripes=True)
    ws.add_table(tab)

    for row in ws.iter_rows(min_row=2, min_col=4, max_col=4, max_row=ws.max_row):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.endswith('%'):
                cell.number_format = '0.00%'

    adjust_column_width(ws)
    wb.save(output_xlsx)
    messagebox.showinfo("Sucesso", f"Análise concluída. Relatório salvo em {output_xlsx}")

def adjust_column_width(ws):
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

def create_gui():
    global window, select_label, analyze_button, progress_var, progress_label, progress_bar, open_folder_button, canvas, directory
    directory = None
    window = ThemedTk(theme="breeze")
    window.title("PDF Analyzer")
    window.geometry("900x650")
    
    main_frame = ttk.Frame(window, padding="20")
    main_frame.pack(fill='both', expand=True)

    label = ttk.Label(main_frame, text="Selecione um diretório para analisar PDFs:", font=("Segoe UI", 12, "bold"))
    label.pack(pady=10)

    select_button = ttk.Button(main_frame, text="Selecionar Diretório", command=select_directory, width=20)
    select_button.pack(pady=10)

    select_label = ttk.Label(main_frame, text="Nenhum diretório selecionado", foreground="gray", font=("Segoe UI", 10))
    select_label.pack(pady=5)

    analyze_button = ttk.Button(main_frame, text="Iniciar Análise", state="disabled", command=start_analysis, width=20)
    analyze_button.pack(pady=15)

    progress_var = StringVar()
    progress_var.set("0")
    progress_label = ttk.Label(main_frame, text="Progresso: 0%", font=("Segoe UI", 10))
    progress_label.pack(pady=10)
    progress_bar = ttk.Progressbar(main_frame, orient="horizontal", length=500, mode="determinate", variable=progress_var)
    progress_bar.pack(pady=10)

    open_folder_button = ttk.Button(main_frame, text="Abrir Pasta do Relatório", command=open_folder, width=20)
    open_folder_button.pack(pady=15)

    canvas_frame = ttk.Frame(main_frame, padding="10", borderwidth=2, relief="ridge")
    canvas_frame.pack(pady=20, expand=True, fill="both")
    
    canvas = Canvas(canvas_frame, bg="white", width=800, height=600)
    canvas.pack(expand=True, fill="both")

    window.mainloop()

if __name__ == "__main__":
    root = Tk()
    root.withdraw()
    test_tesseract_setup()
    root.destroy()
    create_gui()
