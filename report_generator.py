# report_generator.py

import os
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill


class ReportGenerator:
    def __init__(self):
        print("Inicializando ReportGenerator...")
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "PDF Analysis Report"
        self.headers = ["Arquivo PDF", "Página", "Status", "Porcentagem de Pixels Brancos"]
        self.ws.append(self.headers)
        print("Worksheet inicializada com cabeçalhos.")

    def add_record(self, pdf_name, page_num, status, white_pixel_percentage, ocr_performed, extracted_text):
        try:
            print(f"Adicionando registro: PDF Name={pdf_name}, Página={page_num}, Status={status}, Porcentagem de Pixels Brancos={white_pixel_percentage}")
            row = [
                pdf_name,
                page_num,
                status,
                f"{white_pixel_percentage:.2%}"
            ]
            self.ws.append(row)

            # Highlight row in red if the status is "Precisa de Atenção"
            if status == "Precisa de Atenção":
                print("Status 'Precisa de Atenção' detectado. Destacando a linha em vermelho.")
                fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                for col_idx in range(1, len(row) + 1):
                    self.ws.cell(row=self.ws.max_row, column=col_idx).fill = fill

            print("Registro adicionado com sucesso.")
        except Exception as e:
            print(f"Erro ao adicionar registro: {e}")

    def finalize(self, output_path):
        print("Finalizando o relatório...")

        # Garantir que o diretório de destino existe
        dir_path = os.path.dirname(output_path)
        if not os.path.exists(dir_path):
            try:
                os.makedirs(dir_path)
                print(f"Diretório criado: {dir_path}")
            except OSError as e:
                print(f"Erro ao criar o diretório: {e}")
                return

        # Estilo da tabela e ajuste das colunas
        try:
            max_row = self.ws.max_row
            max_col = self.ws.max_column
            table_ref = f"A1:{get_column_letter(max_col)}{max_row}"
            print(f"Criando tabela com referência: {table_ref}")
            tab = Table(displayName="PDFAnalysisTable", ref=table_ref)
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=True
            )
            tab.tableStyleInfo = style
            self.ws.add_table(tab)
            print("Tabela criada com estilo aplicado.")

            for col in self.ws.columns:
                max_length = 0
                column = col[0].column
                column_letter = get_column_letter(column)
                print(f"Ajustando largura da coluna {column_letter}...")
                for cell in col:
                    max_length = max(max_length, len(str(cell.value)))
                adjusted_width = max_length + 2
                self.ws.column_dimensions[column_letter].width = adjusted_width
                print(f"Largura da coluna {column_letter} ajustada para {adjusted_width}.")

            # Salvar o relatório
            self.wb.save(output_path)
            print(f"Relatório salvo em: {output_path}")

        except Exception as e:
            print(f"Erro ao salvar o relatório: {e}")